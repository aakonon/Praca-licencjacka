import re
import random
import requests
from openpyxl import Workbook
from bs4 import BeautifulSoup
import time

def get_eur_pln():
    response = requests.get('https://api.nbp.pl/api/exchangerates/rates/a/eur/?format=json')
    if response.status_code == 200:
        data = response.json()
        return data["rates"][0]["mid"]
    else:
        print('Nie udało się pobrać kursu EUR')
        return 4.2


eur_pln = get_eur_pln()

headers = {
    "X-Requested-With": "XMLHttpRequest",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}

header = ['Adres', 'Cena', 'Powierzchnia',
          'Cena za m2', 'Piętro', 'Liczba pokoi', 'Link']


book = Workbook()
page = book.active
page.append(header)
row = 2

url = input('Link do strony: \n')
data = requests.get(f'{url}&ajax=1', headers=headers).json()

typ = url.split(',')[2].title()
miasto = data['searchValues']['cityName'].title()
ostatnia_strona = data['totalPages']
hrefs = set()

for p in random.sample(range(1, ostatnia_strona + 1), k=ostatnia_strona):
    if len(hrefs) >= 300:
        break
    time.sleep(1.5)
    response = requests.get(f'{url}&p={p}&ajax=1', headers=headers)
    print(response.url)
    data = response.json()
    src = data['listHTML']
    soup = BeautifulSoup(src, "lxml")
    ogloszenia = data['listAdditionalData']
    for k, j in ogloszenia.items():
        is_archieved = j['asArchive']
        is_supplement = j['isSupplementAd']
        href = j['shareUrl']
        if not is_archieved and not is_supplement and href not in hrefs:
            hrefs.add(href)
            page.cell(row, 1, miasto)
            cena = float(j['primaryPrice'])
            ogloszenie = soup.find(attrs={'data-id': f'{k}'})
            if ogloszenie.find(string=lambda s: s and '€' in s):
                cena = cena * eur_pln
            if cena:
                page.cell(row, 2, cena)
            pattern = re.compile(r'\b\d{1,3}[,.]?\d{0,2}\s?m[²2]\b', re.IGNORECASE)
            m2 = ogloszenie.find(string=pattern)
            if m2:
                m2 = m2.replace('m²', '').replace(',', '.').strip()
                page.cell(row, 3, float(m2))
            if cena and m2:
                za_m2 = float(cena) / float(m2)
                page.cell(row, 4, round(za_m2, 2))

            attrs = {i['label']: i['values'][0]['value'] for i in j['attributes']}
            page.cell(row, 5, attrs["Piętro"])
            page.cell(row, 6, attrs['Liczba pokoi'])
            page.cell(row, 7).hyperlink = href

            row += 1

book.save(f'{typ}/{miasto}.xlsx')
